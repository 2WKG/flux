"""Cut the committed EIA-860 test slice out of the real 2025 Early Release archive.

The slice is a byte-reduced copy of the published archive, not a synthetic
workbook: every retained cell -- the two preamble rows, the header row, and each
data row -- is copied verbatim from the source workbook.  Only whole rows for a
small set of plants and whole non-selected sheets' data rows are dropped, so the
zip member names, sheet names, header offset and column set stay exactly as EIA
published them.  ``pipelines/tests/test_eia860_physical.py`` runs the real
parser over this slice with no ``pd.read_excel`` patch.

Re-derive with::

    curl -L -o eia8602025ER.zip \
        https://www.eia.gov/electricity/data/eia860/xls/eia8602025ER.zip
    shasum -a 256 eia8602025ER.zip   # must equal SOURCE_SHA256 below
    uv run python scripts/cut_eia860_test_slice.py eia8602025ER.zip \
        pipelines/tests/fixtures/eia860/eia8602025ER-slice.zip
"""

from __future__ import annotations

import argparse
import hashlib
import io
import re
import sys
import zipfile
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pandas as pd

# Make ``pipelines`` importable when this file is run as a script from any cwd.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from pipelines.eia860_physical import (
    _HEADER_ROW,
    GENERATOR_MEMBER,
    PLANT_MEMBER,
    STORAGE_MEMBER,
)

SOURCE_URL = "https://www.eia.gov/electricity/data/eia860/xls/eia8602025ER.zip"
SOURCE_SHA256 = "bd05bac5149371a6aab869e8a4c737ccb2e6c2f83f0daadac895982af9510bf1"

# Plants chosen so the slice exercises every branch the parser has: a plant with
# operable, retired and storage rows (8063), a plant whose Schedule 2 coordinate
# cells are blank (7732), a plant with proposed rows (69414), and two Minnesota
# plants (62908 operable storage, 2038 proposed storage).
SELECTED_PLANT_CODES = (2038, 7732, 8063, 62908, 69414)

# The slice is byte-reproducible: workbook timestamps and zip member dates are
# pinned so re-running this script yields the identical sha256.
_EPOCH = datetime(1980, 1, 1, tzinfo=UTC)

_UNIT_SHEETS = ("Operable", "Proposed", "Retired and Canceled")
_MEMBER_SHEETS = {
    PLANT_MEMBER: ("Plant",),
    GENERATOR_MEMBER: _UNIT_SHEETS,
    STORAGE_MEMBER: _UNIT_SHEETS,
}


def _plant_code(value: object) -> int | None:
    try:
        number = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else None


def _slice_member(payload: bytes, sheets: tuple[str, ...]) -> bytes:
    source = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    target = openpyxl.Workbook()
    target.remove(target.active)
    for sheet_name in sheets:
        rows = source[sheet_name].iter_rows(values_only=True)
        out = target.create_sheet(sheet_name)
        header: list[object] = []
        for index, row in enumerate(rows):
            if index <= _HEADER_ROW:
                # Preamble and header rows are copied verbatim, so _HEADER_ROW
                # and the published column names stay under test.
                out.append(list(row))
                header = list(row)
                continue
            code_index = header.index("Plant Code")
            if _plant_code(row[code_index]) in SELECTED_PLANT_CODES:
                out.append(list(row))
    source.close()
    target.properties.created = _EPOCH
    target.properties.modified = _EPOCH
    target.properties.creator = ""
    target.properties.lastModifiedBy = ""
    buffer = io.BytesIO()
    target.save(buffer)
    # openpyxl stamps each zip member with the current clock; repin them.
    return _repin_zip_dates(buffer.getvalue())


def _repin_zip_dates(payload: bytes) -> bytes:
    buffer = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload)) as source,
        zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as destination,
    ):
        for info in source.infolist():
            pinned = zipfile.ZipInfo(info.filename, date_time=_EPOCH.timetuple()[:6])
            pinned.compress_type = zipfile.ZIP_DEFLATED
            pinned.external_attr = info.external_attr
            content = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                # openpyxl rewrites dcterms:modified from the clock on save.
                content = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]+(</dcterms:modified>)",
                    rb"\g<1>"
                    + _EPOCH.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
                    + rb"\g<2>",
                    content,
                )
            destination.writestr(pinned, content)
    return buffer.getvalue()


def cut_slice(archive_path: Path, output_path: Path) -> Path:
    digest = hashlib.sha256(archive_path.read_bytes()).hexdigest()
    if digest != SOURCE_SHA256:
        raise SystemExit(
            f"{archive_path} sha256 {digest} does not match the receipted "
            f"{SOURCE_SHA256}; refusing to cut a slice from an unverified archive"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with (
        zipfile.ZipFile(archive_path) as source,
        zipfile.ZipFile(
            output_path, "w", compression=zipfile.ZIP_DEFLATED
        ) as destination,
    ):
        for member, sheets in _MEMBER_SHEETS.items():
            info = zipfile.ZipInfo(member, date_time=_EPOCH.timetuple()[:6])
            info.compress_type = zipfile.ZIP_DEFLATED
            destination.writestr(info, _slice_member(source.read(member), sheets))
    return output_path


def _report(output_path: Path) -> None:
    with zipfile.ZipFile(output_path) as archive:
        for member, sheets in _MEMBER_SHEETS.items():
            for sheet in sheets:
                frame = pd.read_excel(
                    io.BytesIO(archive.read(member)),
                    sheet_name=sheet,
                    header=_HEADER_ROW,
                )
                print(f"{member}:{sheet} rows={len(frame)} cols={len(frame.columns)}")
    print(
        f"{output_path} bytes={output_path.stat().st_size} "
        f"sha256={hashlib.sha256(output_path.read_bytes()).hexdigest()}"
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("archive", type=Path, help="the real eia8602025ER.zip")
    parser.add_argument("output", type=Path, help="slice zip to write")
    args = parser.parse_args(argv)
    _report(cut_slice(args.archive, args.output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
